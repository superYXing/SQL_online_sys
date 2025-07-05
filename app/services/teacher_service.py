from typing import Optional, List, Dict, Tuple, Any
from sqlalchemy.orm import Session
from sqlalchemy import func, case, distinct, and_
from models import (
    Teacher, Course, Semester, Student, CourseSelection, AnswerRecord, Problem,
    DatabaseSchema
)
from schemas.teacher import (
    TeacherProfileResponse, StudentCreateRequest,
    StudentCreateResponse, ImportFailDetail, StudentImportResponse,
    ScoreCalculateRequest, StudentScoreInfo, ScoreUpdateResponse, ScoreListResponse,
    TeacherStudentInfo, TeacherStudentListResponse, ScoreExportRequest, ExportStudentInfo,
    SQLQueryResponse, StudentProfileDocResponse,
    DatasetExportResponse, StudentCourseAddRequest, StudentCourseAddResponse,
    SchemaCreateRequest, SchemaCreateResponse, SQLQueryRequest, SQLQueryResponse
)
# 已删除无用导入: CourseInfo, TeacherCourseListResponse, StudentGradeInfo, CourseGradeResponse, ProblemStatisticsResponse, DashboardMatrixResponse
from datetime import datetime
import pandas as pd
import io
import json
from services.public_service import public_service

class TeacherService:
    """教师服务类"""
    
    def get_teacher_profile(self, teacher_id: str, db: Session) -> Optional[TeacherProfileResponse]:
        """获取教师个人信息"""
        try:
            # 获取教师基本信息
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return None

            # 获取当前学期（调用公共服务）
            from services.public_service import public_service
            current_semester = public_service.get_current_semester(db)
            semester_name = current_semester.semester_name if current_semester else "暂无学期"

            # 获取教师信息，这里假设任课教师就是教师本人
            # 根据实际业务逻辑调整
            return TeacherProfileResponse(
                teacher_id=teacher.teacher_id,
                teacher_name=teacher.teacher_name or "未知",
                semester_name=semester_name,
            )

        except Exception as e:
            print(f"获取教师个人信息失败: {e}")
            return None

    # 已删除: get_teacher_courses 方法 - 功能已整合到其他方法

    # 已删除: get_course_grades 方法 - 功能已整合到其他方法

    # 已删除: export_course_grades 方法 - 功能已整合到其他方法

    def create_student(self, teacher_id: str, student_data: StudentCreateRequest,
                      db: Session) -> Tuple[bool, str, Optional[StudentCreateResponse]]:
        """教师创建学生"""
        try:
            # 验证教师是否存在
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return False, "教师不存在", None

            # 检查学生ID是否已存在
            existing_student = db.query(Student).filter(
                Student.student_id == student_data.student_id
            ).first()
            if existing_student:
                return False, f"学生ID {student_data.student_id} 已存在", None

            # 创建新学生
            new_student = Student(
                student_id=student_data.student_id,
                student_name=student_data.student_name,
                class_=student_data.class_,
                student_password=student_data.student_password
            )

            db.add(new_student)
            db.commit()
            db.refresh(new_student)

            # 构建响应
            response = StudentCreateResponse(
                id=new_student.id,
                student_id=new_student.student_id,
                student_name=new_student.student_name,
                class_=new_student.class_
            )

            return True, "学生创建成功", response

        except Exception as e:
            db.rollback()
            print(f"创建学生失败: {e}")
            return False, f"创建失败: {str(e)}", None

    def add_student_course(self, teacher_id: str, course_data: StudentCourseAddRequest,
                          db: Session) -> Tuple[bool, str, Optional[StudentCourseAddResponse]]:
        """添加学生选课信息"""
        try:
            # 验证教师是否存在
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return False, "教师不存在", None

            # 验证课程是否存在
            course = db.query(Course).filter(Course.course_id == course_data.course_id).first()
            if not course:
                return False, f"课程ID {course_data.course_id} 不存在", None

            # 获取当前学期ID
            current_semester = public_service.get_current_semester(db)
            if not current_semester:
                return False, "无法获取当前学期信息", None

            # 检查学生是否已存在
            existing_student = db.query(Student).filter(
                Student.student_id == course_data.student_id
            ).first()

            if existing_student:
                # 学生已存在，检查是否已经选择了该课程
                existing_selection = db.query(CourseSelection).filter(
                    CourseSelection.student_id == existing_student.id,
                    CourseSelection.course_id == course_data.course_id,
                    CourseSelection.semester_id == current_semester.semester_id
                ).first()

                if existing_selection:
                    return False, f"学生 {course_data.student_id} 已经选择了课程 {course_data.course_id}", None

                # 创建选课记录
                new_selection = CourseSelection(
                    student_id=existing_student.id,
                    course_id=course_data.course_id,
                    status=course_data.status,
                    semester_id=current_semester.semester_id,
                    score=0  # 默认分数为0
                )

                db.add(new_selection)
                db.commit()

                response = StudentCourseAddResponse(code=200, msg="添加")
                return True, "学生选课信息添加成功", response

            else:
                # 学生不存在，先创建学生
                new_student = Student(
                    student_id=course_data.student_id,
                    student_name=course_data.student_name,
                    class_=course_data.class_,
                    student_password="default@password"  # 默认密码
                )

                db.add(new_student)
                db.flush()  # 获取新学生的ID

                # 创建选课记录
                new_selection = CourseSelection(
                    student_id=new_student.id,
                    course_id=course_data.course_id,
                    status=course_data.status,
                    semester_id=current_semester.semester_id,
                    score=0  # 默认分数为0
                )

                db.add(new_selection)
                db.commit()

                response = StudentCourseAddResponse(code=200, msg="添加")
                return True, "学生创建并选课成功", response

        except Exception as e:
            db.rollback()
            print(f"添加学生选课信息失败: {e}")
            return False, f"添加失败: {str(e)}", None

    def import_students_from_excel(self, teacher_id: str, file_content: bytes,
                                  db: Session) -> StudentImportResponse:
        """从Excel文件批量导入学生"""
        try:
            # 验证教师是否存在
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return StudentImportResponse(
                    code=400,
                    msg="教师不存在",
                    success_count=0,
                    fail_count=0,
                    fail_details=[]
                )

            # 读取Excel文件
            try:
                df = pd.read_excel(io.BytesIO(file_content))
            except Exception as e:
                return StudentImportResponse(
                    code=400,
                    msg=f"Excel文件格式错误: {str(e)}",
                    success_count=0,
                    fail_count=0,
                    fail_details=[]
                )

            # 验证必需的列
            required_columns = ['学号', '姓名', '班级', '密码']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return StudentImportResponse(
                    code=400,
                    msg=f"Excel文件缺少必需的列: {', '.join(missing_columns)}",
                    success_count=0,
                    fail_count=0,
                    fail_details=[]
                )

            success_count = 0
            fail_count = 0
            fail_details = []

            # 逐行处理学生数据
            for index, row in df.iterrows():
                try:
                    student_id = str(row['学号']).strip()
                    student_name = str(row['姓名']).strip()
                    class_ = str(row['班级']).strip()
                    password = str(row['密码']).strip()

                    # 验证数据
                    if not student_id or not student_name or not class_ or not password:
                        fail_count += 1
                        fail_details.append(ImportFailDetail(
                            row=index + 2,  # Excel行号从2开始（包含表头）
                            reason="数据不完整，请检查学号、姓名、班级、密码是否都已填写"
                        ))
                        continue

                    # 检查学生ID是否已存在
                    existing_student = db.query(Student).filter(
                        Student.student_id == student_id
                    ).first()
                    if existing_student:
                        fail_count += 1
                        fail_details.append(ImportFailDetail(
                            row=index + 2,
                            reason=f"学生ID {student_id} 已存在"
                        ))
                        continue

                    # 创建学生
                    new_student = Student(
                        student_id=student_id,
                        student_name=student_name,
                        class_=class_,
                        student_password=password
                    )

                    db.add(new_student)
                    success_count += 1

                except Exception as e:
                    fail_count += 1
                    fail_details.append(ImportFailDetail(
                        row=index + 2,
                        reason=f"处理数据时出错: {str(e)}"
                    ))

            # 提交事务
            if success_count > 0:
                db.commit()

            return StudentImportResponse(
                code=200,
                msg=f"导入完成，成功 {success_count} 条，失败 {fail_count} 条",
                success_count=success_count,
                fail_count=fail_count,
                fail_details=fail_details
            )

        except Exception as e:
            db.rollback()
            print(f"批量导入学生失败: {e}")
            return StudentImportResponse(
                code=500,
                msg=f"导入失败: {str(e)}",
                success_count=0,
                fail_count=0,
                fail_details=[]
            )

    def calculate_scores(self, teacher_id: str, problem_ids: List[int],
                        db: Session) -> ScoreUpdateResponse:
        """教师核算分数"""
        try:
            # 验证教师是否存在
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return ScoreUpdateResponse(
                    code=400,
                    msg="教师不存在"
                )

            # 获取当前学期
            from services.public_service import public_service
            current_semester = public_service.get_current_semester(db)
            if not current_semester:
                return ScoreUpdateResponse(
                    code=400,
                    msg="未找到当前学期"
                )

            # 获取本学期的所有课程
            courses = db.query(Course).filter(
                Course.semester_id == current_semester.semester_id
            ).all()

            if not courses:
                return ScoreUpdateResponse(
                    code=400,
                    msg="当前学期没有课程"
                )

            score_list = []

            # 遍历每门课程
            for course in courses:
                # 获取选课学生及其选课记录
                course_selections = db.query(
                    CourseSelection,
                    Student.student_id,
                    Student.student_name,
                    Student.class_
                ).join(
                    Student, CourseSelection.student_id == Student.id
                ).filter(
                    CourseSelection.course_id == course.course_id
                ).all()

                # 计算每个学生的分数
                for course_selection, student_id, student_name, student_class in course_selections:
                    # 查询学生在指定题目上的正确答题数（result_type == 0表示正确）
                    correct_count = db.query(func.count(distinct(AnswerRecord.problem_id))).filter(
                        AnswerRecord.student_id == course_selection.student_id,
                        AnswerRecord.problem_id.in_(problem_ids),
                        AnswerRecord.result_type == 0
                    ).scalar() or 0

                    # 每道题10分
                    total_score = correct_count * 10
                    # 分数上限是100分
                    total_score = min(total_score, 100)

                    # 更新选课记录中的分数
                    course_selection.score = total_score

                    score_list.append(StudentScoreInfo(
                        course_id=course.course_id,
                        student_id=student_id,
                        student_name=student_name or "未知",
                        class_=student_class or "",
                        total_score=total_score
                    ))

            # 提交数据库事务，保存分数更新
            db.commit()

            return ScoreUpdateResponse(
                code=200,
                msg="更新分数成功"
            )

        except Exception as e:
            print(f"核算分数失败: {e}")
            return ScoreUpdateResponse(
                code=500,
                msg=f"核算失败: {str(e)}"
            )

    def get_student_scores(self, teacher_id: str, db: Session) -> ScoreListResponse:
        """获取学生分数列表"""
        try:
            # 验证教师是否存在
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return ScoreListResponse(
                    code=400,
                    msg="教师不存在",
                    scorelist=[]
                )

            # 获取当前学期
            from services.public_service import public_service
            current_semester = public_service.get_current_semester(db)
            if not current_semester:
                return ScoreListResponse(
                    code=400,
                    msg="未找到当前学期",
                    scorelist=[]
                )

            # 获取当前学期的所有课程
            courses = db.query(Course).filter(
                Course.semester_id == current_semester.semester_id
            ).all()

            if not courses:
                return ScoreListResponse(
                    code=200,
                    msg="",
                    scorelist=[]
                )

            score_list = []

            # 遍历每门课程，获取学生分数
            for course in courses:
                # 获取选课学生及其分数
                course_selections = db.query(
                    CourseSelection,
                    Student.student_id,
                    Student.student_name,
                    Student.class_
                ).join(
                    Student, CourseSelection.student_id == Student.id
                ).filter(
                    CourseSelection.course_id == course.course_id
                ).all()

                # 构建分数列表
                for course_selection, student_id, student_name, student_class in course_selections:
                    score_list.append(StudentScoreInfo(
                        course_id=course.course_id,
                        student_id=student_id,
                        student_name=student_name or "未知",
                        class_=student_class or "",
                        total_score=course_selection.score or 0
                    ))

            return ScoreListResponse(
                code=200,
                msg="",
                scorelist=score_list
            )

        except Exception as e:
            print(f"获取学生分数失败: {e}")
            return ScoreListResponse(
                code=500,
                msg=f"获取失败: {str(e)}",
                scorelist=[]
            )

    def export_scores(self, teacher_id: str, course_id: int, class_id: Optional[str] = None,
                     db: Session = None) -> Optional[Dict]:
        """导出分数结果"""
        try:
            # 验证教师权限
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return None

            course = db.query(Course).filter(
                Course.course_id == course_id,
                Course.teacher_id == teacher.id
            ).first()
            if not course:
                return None

            # 获取当前学期
            from services.public_service import public_service
            current_semester = public_service.get_current_semester(db)
            semester_name = current_semester.semester_name if current_semester else "未知学期"

            # 构建查询条件
            query = db.query(
                Student.student_id,
                Student.student_name,
                Student.class_
            ).select_from(CourseSelection).join(
                Student, CourseSelection.student_id == Student.id
            ).filter(
                CourseSelection.course_id == course_id
            )

            # 如果指定了班级，添加班级过滤
            if class_id:
                query = query.filter(Student.class_ == class_id)

            students = query.all()

            # 构建导出数据
            export_data = {
                "course_info": {
                    "course_id": course_id,
                    "course_name": course.course_name or "未命名课程",
                    "semester_name": semester_name,
                    "class_filter": class_id or "全部班级",
                    "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "total_students": len(students)
                },
                "students": []
            }
            for student in students:
                # 根据学号和课程号查询选课表里的分数
                query = db.query(
                    CourseSelection.score
                ).select_from(CourseSelection).join(
                    Student, CourseSelection.student_id == Student.id
                ).filter(
                    CourseSelection.course_id == course_id,
                    CourseSelection.student_id == student.student_id

                )
                export_data["students"].append({
                    "学号": student.student_id,
                    "姓名": student.student_name,
                    "班级": student.class_,
                    "总分": CourseSelection.score  # 模拟分数
                })

            return export_data

        except Exception as e:
            print(f"导出分数失败: {e}")
            return None

    def get_students_by_semester(self, page: int = 1, limit: int = 20, search: Optional[str] = None,
                                class_filter: Optional[str] = None, semester_id: Optional[int] = None,
                                db: Session = None) -> TeacherStudentListResponse:
        """根据学期获取学生列表"""
        try:
            from sqlalchemy import or_

            # 构建基础查询
            query = db.query(
                Student.student_id,
                Student.student_name,
                Student.class_,
                Teacher.teacher_name,
                Semester.semester_name,
                Course.course_id
            ).select_from(CourseSelection).join(
                Student, CourseSelection.student_id == Student.id
            ).join(
                Course, CourseSelection.course_id == Course.course_id
            ).join(
                Teacher, Course.teacher_id == Teacher.id
            ).join(
                Semester, Course.semester_id == Semester.semester_id
            )

            # 如果指定了学期ID，则过滤
            if semester_id is not None:
                query = query.filter(Semester.semester_id == semester_id)

            # 搜索功能
            if search:
                query = query.filter(
                    or_(
                        Student.student_id.contains(search),
                        Student.student_name.contains(search)
                    )
                )

            # 班级过滤
            if class_filter:
                query = query.filter(Student.class_.contains(class_filter))

            # 获取总数
            total = query.count()

            # 分页查询
            offset = (page - 1) * limit
            results = query.offset(offset).limit(limit).all()

            # 构建响应数据
            student_list = []
            for result in results:
                student_list.append(TeacherStudentInfo(
                    student_id=result.student_id,
                    student_name=result.student_name or "",
                    class_=result.class_ or "",
                    teacher_name=result.teacher_name or "",
                    semester_name=result.semester_name or "",
                    course_id=str(result.course_id) if result.course_id else ""
                ))

            return TeacherStudentListResponse(
                students=student_list,
                total=total,
                page=page,
                limit=limit
            )

        except Exception as e:
            print(f"获取学生列表失败: {e}")
            return TeacherStudentListResponse(students=[], total=0, page=page, limit=limit)

    def export_scores_to_excel(self, students_data: List[ExportStudentInfo]) -> bytes:
        """将学生分数数据导出为Excel文件"""
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from datetime import datetime

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "成绩表"

            # 设置表头
            headers = ["学号", "姓名", "班级", "课序号", "状态", "总分数"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 填充数据
            for row, student in enumerate(students_data, 2):
                ws.cell(row=row, column=1, value=student.student_id)
                ws.cell(row=row, column=2, value=student.student_name)
                ws.cell(row=row, column=3, value=student.class_)
                ws.cell(row=row, column=4, value=student.course_id)
                ws.cell(row=row, column=5, value=student.status)
                ws.cell(row=row, column=6, value=student.total_score)

                # 设置数据行样式
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # 调整列宽
            column_widths = [15, 10, 15, 8, 8, 10]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width

            # 保存到内存
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            return excel_buffer.getvalue()

        except Exception as e:
            print(f"生成Excel文件失败: {e}")
            raise Exception(f"生成Excel文件失败: {str(e)}")

    def generate_export_filename(self, students_data: List[ExportStudentInfo]) -> str:
        """生成导出文件名"""
        try:
            from datetime import datetime

            if not students_data:
                return f"成绩表-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # 获取学期和班级信息（这里简化处理，实际可以从数据库查询）
            first_student = students_data[0]
            class_name = first_student.class_

            # 提取班级中的数字部分作为班级标识
            import re
            class_match = re.search(r'(\d+)', class_name)
            class_num = class_match.group(1) if class_match else "未知班级"

            # 生成文件名：学期-班级-成绩表.xlsx
            current_year = datetime.now().year
            semester = f"{current_year}年春季"  # 可以根据实际情况调整
            filename = f"{semester}-{class_num}班-成绩表.xlsx"

            return filename

        except Exception as e:
            print(f"生成文件名失败: {e}")
            return f"成绩表-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 已删除: get_dashboard_matrix 方法 - 接口已废弃

    def execute_sql_query(self, teacher_id: str, query_data: SQLQueryRequest, db: Session) -> SQLQueryResponse:
        """执行SQL查询"""
        try:
            # 验证教师是否存在
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return SQLQueryResponse(
                    code=400,
                    msg="教师不存在",
                    columns=[],
                    rows=[]
                )

            # 验证数据库模式是否存在
            schema = db.query(DatabaseSchema).filter(DatabaseSchema.schema_id == query_data.schema_id).first()
            if not schema:
                return SQLQueryResponse(
                    code=400,
                    msg="数据库模式不存在",
                    columns=[],
                    rows=[]
                )

            # 使用database_engine_service执行SQL查询，默认使用PostgreSQL
            from services.database_engine_service import database_engine_service

            # 构建完整的SQL语句，包含数据库切换语句
            if schema.sql_schema:
                # 根据引擎类型选择切换语句（默认使用PostgreSQL）
                engine_type = "postgresql"
                if engine_type == "mysql":
                    switch_sql = f"USE {schema.sql_schema};"
                elif engine_type in ["postgresql", "opengauss"]:
                    switch_sql = f"SET search_path TO {schema.sql_schema};"
                else:
                    switch_sql = f"SET search_path TO {schema.sql_schema};"  # 默认使用PostgreSQL语法
                # 输出切换语句
                print(f"执行切换语句: {switch_sql}")
                # 执行切换语句
                switch_success, switch_message, _ = database_engine_service.execute_sql(
                    sql=switch_sql,
                    engine_type=engine_type
                )

                if not switch_success:
                    print(f"执行数据库切换失败: {switch_message}")
                    # 继续执行用户SQL，不因切换失败而中断

            # 执行用户的SQL查询
            success, message, result_data = database_engine_service.execute_sql(
                sql=query_data.sql,
                engine_type="postgresql"
            )

            if not success:
                return SQLQueryResponse(
                    code=400,
                    msg=f"SQL执行失败: {message}",
                    columns=[],
                    rows=[]
                )

            # 处理查询结果
            if result_data and len(result_data) > 0:
                # 获取列名
                columns = list(result_data[0].keys()) if result_data else []

                # 将字典列表转换为二维数组
                rows = []
                for row_dict in result_data:
                    row = [row_dict.get(col) for col in columns]
                    rows.append(row)

                return SQLQueryResponse(
                    code=200,
                    msg="查询成功",
                    columns=columns,
                    rows=rows
                )
            else:
                # 空结果集
                return SQLQueryResponse(
                    code=200,
                    msg="查询成功",
                    columns=[],
                    rows=[]
                )

        except Exception as e:
            print(f"执行SQL查询失败: {e}")
            return SQLQueryResponse(
                code=500,
                msg=f"查询执行失败: {str(e)}",
                columns=[],
                rows=[]
            )

    def get_student_profile_doc(self, student_id: str, schema_id: int, db: Session) -> Optional[StudentProfileDocResponse]:
        """获取学生答题概况"""
        try:
            # 1. 调用public接口获取当前学期号
            from services.public_service import public_service
            current_semester = public_service.get_current_semester(db)
            if not current_semester:
                return None

            current_semester_id = current_semester.semester_id

            # 2. 根据学期号确认课程号范围
            courses = db.query(Course).filter(Course.semester_id == current_semester_id).all()
            if not courses:
                return None

            # 3. 根据学号在课程号范围里确认课程号(使用.first())
            # 获取学生信息
            student = db.query(Student).filter(Student.student_id == student_id).first()
            if not student:
                return None

            # 查找学生在当前学期的选课记录
            course_selection = db.query(CourseSelection).join(
                Course, CourseSelection.course_id == Course.course_id
            ).filter(
                CourseSelection.student_id == student.id,
                Course.semester_id == current_semester_id
            ).first()

            if not course_selection:
                return None

            course_id = course_selection.course_id
            status = course_selection.status if hasattr(course_selection, 'status') else 1  # 默认为正常

            # 4. 根据数据库模式id得到题目id范围
            problems = db.query(Problem).filter(Problem.schema_id == schema_id).all()
            if not problems:
                # 如果没有题目，返回基本信息但答题统计为0
                return StudentProfileDocResponse(
                    student_id=student.student_id,
                    student_name=student.student_name or "",
                    class_name=student.class_ or "",
                    course_id=course_id,
                    status=status,
                    correct_count=0,
                    submit_count=0
                )

            problem_ids = [p.problem_id for p in problems]

            # 5. 到答题记录表里根据学号和题目的范围确认答题的正确数量和总提交数
            answer_records = db.query(AnswerRecord).filter(
                AnswerRecord.student_id == student.id,
                AnswerRecord.problem_id.in_(problem_ids)
            ).all()

            # 计算正确数量和总提交数
            submit_count = len(answer_records)
            correct_count = sum(1 for record in answer_records if record.result_type == 0)

            return StudentProfileDocResponse(
                student_id=student.student_id,
                student_name=student.student_name or "",
                class_name=student.class_ or "",
                course_id=course_id,
                status=status,
                correct_count=correct_count,
                submit_count=submit_count
            )

        except Exception as e:
            print(f"获取学生答题概况失败: {e}")
            return None

    def export_dataset(self, schema_name: str, format: str, db: Session) -> Tuple[Optional[Any], str, str]:
        """导出数据库模式相关数据"""
        try:
            # 获取数据库模式
            schema = db.query(DatabaseSchema).filter(DatabaseSchema.schema_name == schema_name).first()
            if not schema:
                return None, "数据库模式不存在", ""

            # 获取当前学期
            from services.public_service import public_service
            current_semester = public_service.get_current_semester(db)
            if not current_semester:
                return None, "未找到当前学期", ""

            # 根据数据库模式获取题目范围
            problems = db.query(Problem).filter(Problem.schema_id == schema.schema_id).all()
            if not problems:
                return None, "该数据库模式下没有题目", ""

            problem_ids = [p.problem_id for p in problems]

            # 获取当前学期的所有学生
            students = db.query(Student).join(
                CourseSelection, Student.id == CourseSelection.student_id
            ).join(
                Course, CourseSelection.course_id == Course.course_id
            ).filter(
                Course.semester_id == current_semester.semester_id
            ).distinct().all()

            # 构建导出数据
            export_data = []
            for student in students:
                # 查询该学生在指定题目范围内的所有提交记录
                answer_records = db.query(AnswerRecord).filter(
                    AnswerRecord.student_id == student.id,
                    AnswerRecord.problem_id.in_(problem_ids)
                ).all()

                if answer_records:  # 只导出有提交记录的学生
                    # 计算统计数据
                    submit_count = len(answer_records)
                    correct_count = sum(1 for record in answer_records if record.result_type == 0)

                    # 计算完成题目总数（去重）
                    completed_problem_ids = set(record.problem_id for record in answer_records)
                    problem_count = len(completed_problem_ids)

                    # 计算不同解法数量（根据不同的SQL语句）
                    unique_sqls = set(record.student_answer for record in answer_records if record.student_answer)
                    method_count = len(unique_sqls)

                    export_data.append({
                        "student_id": student.student_id,
                        "student_name": student.student_name or "",
                        "class": student.class_ or "",
                        "problem_count": problem_count,
                        "submit_count": submit_count,
                        "correct_count": correct_count,
                        "method_count": method_count
                    })

            if not export_data:
                return None, "没有找到相关数据", ""

            # 根据格式生成文件
            if format.upper() == "XLSX":
                file_data, filename, media_type = self._generate_excel_export(export_data, schema_name)
                return file_data, "", media_type
            elif format.upper() == "JSON":
                file_data, filename, media_type = self._generate_json_export(export_data, schema_name)
                return file_data, "", media_type
            elif format.upper() == "XML":
                file_data, filename, media_type = self._generate_xml_export(export_data, schema_name)
                return file_data, "", media_type
            else:
                return None, "不支持的导出格式", ""

        except Exception as e:
            print(f"导出数据失败: {e}")
            return None, f"导出数据失败: {str(e)}", ""

    def _generate_excel_export(self, data: List[Dict], schema_name: str) -> Tuple[io.BytesIO, str, str]:
        """生成Excel格式导出"""
        try:
            df = pd.DataFrame(data)

            # 重命名列
            df.columns = ["学号", "姓名", "班级", "题目数", "提交数", "正确数", "解法数"]

            # 创建Excel文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=schema_name, index=False)

            output.seek(0)
            filename = f"{schema_name}_数据导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            return output, filename, media_type

        except Exception as e:
            print(f"生成Excel导出失败: {e}")
            raise

    def _generate_json_export(self, data: List[Dict], schema_name: str) -> Tuple[io.BytesIO, str, str]:
        """生成JSON格式导出"""
        try:
            json_str = json.dumps(data, ensure_ascii=False, indent=2)
            output = io.BytesIO(json_str.encode('utf-8'))
            filename = f"{schema_name}_数据导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            media_type = "application/json"

            return output, filename, media_type

        except Exception as e:
            print(f"生成JSON导出失败: {e}")
            raise

    def _generate_xml_export(self, data: List[Dict], schema_name: str) -> Tuple[io.BytesIO, str, str]:
        """生成XML格式导出"""
        try:
            # 简单的XML生成
            xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>']
            xml_lines.append(f'<{schema_name}_export>')

            for item in data:
                xml_lines.append('  <student>')
                for key, value in item.items():
                    xml_lines.append(f'    <{key}>{value}</{key}>')
                xml_lines.append('  </student>')

            xml_lines.append(f'</{schema_name}_export>')

            xml_str = '\n'.join(xml_lines)
            output = io.BytesIO(xml_str.encode('utf-8'))
            filename = f"{schema_name}_数据导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
            media_type = "application/xml"

            return output, filename, media_type

        except Exception as e:
            print(f"生成XML导出失败: {e}")
            raise

    # 已删除: get_problem_statistics 方法 - 功能已整合到其他方法

    def create_database_schema(self, teacher_id: str, schema_data: SchemaCreateRequest,
                              db: Session) -> Tuple[bool, str, Optional[SchemaCreateResponse]]:
        """根据HTML格式文本、数据库模式名称、SQL引擎和SQL建表文件创建数据库模式"""
        try:
            # 验证教师是否存在
            teacher = db.query(Teacher).filter(Teacher.teacher_id == teacher_id).first()
            if not teacher:
                return False, "教师不存在", None

            # 验证必要参数
            if not schema_data.html_content or not schema_data.html_content.strip():
                return False, "HTML内容不能为空", None
            if not schema_data.schema_name or not schema_data.schema_name.strip():
                return False, "数据库模式名称不能为空", None
            if not schema_data.sql_engine or not schema_data.sql_engine.strip():
                return False, "SQL引擎类型不能为空", None
            if not schema_data.sql_file_content or not schema_data.sql_file_content.strip():
                return False, "SQL文件内容不能为空", None
            if not schema_data.sql_schema or not schema_data.sql_schema.strip():
                return False, "SQL模式名称不能为空", None

            # 1. 判断此模式是否已经存在
            existing_schema = db.query(DatabaseSchema).filter(
                DatabaseSchema.schema_name == schema_data.schema_name
            ).first()
            if existing_schema:
                return False, f"数据库模式 '{schema_data.schema_name}' 已存在", None

            # 2. 验证SQL引擎类型
            supported_engines = ["mysql", "postgresql", "opengauss"]
            if schema_data.sql_engine.lower() not in supported_engines:
                return False, f"不支持的SQL引擎类型: {schema_data.sql_engine}。支持的类型: {', '.join(supported_engines)}", None

            # 3. 连接指定的SQL引擎并执行SQL文件
            from services.database_engine_service import database_engine_service

            # 执行SQL文件中的建表语句
            success, error_msg, _ = database_engine_service.execute_sql(
                sql=schema_data.sql_file_content,
                engine_type=schema_data.sql_engine.lower()
            )

            if not success:
                return False, f"执行SQL文件失败: {error_msg}", None

            # 4. 执行成功后将数据插入到database_schema表中
            new_schema = DatabaseSchema(
                schema_name=schema_data.schema_name,
                schema_discription=schema_data.html_content,  # 使用HTML内容作为描述
                sql_schema=schema_data.sql_schema,  # 保存SQL模式名称
                schema_author=teacher.teacher_name or teacher.teacher_id  # 设置作者
            )

            db.add(new_schema)
            db.commit()
            db.refresh(new_schema)

            response = SchemaCreateResponse(
                code=200,
                msg="创建数据库模式成功"
            )

            return True, "创建数据库模式成功", response

        except Exception as e:
            db.rollback()
            print(f"创建数据库模式失败: {e}")
            return False, f"创建失败: {str(e)}", None

# 全局教师服务实例
teacher_service = TeacherService()
